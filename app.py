import streamlit as st
import pandas as pd
import cloudinary
import cloudinary.uploader
import cloudinary.api
import hashlib
import requests
import io
import openpyxl
from openpyxl.styles import PatternFill

# --- 1. KONFIGURASI HALAMAN ---
st.set_page_config(page_title="Web Monitoring IC Bali", layout="wide", page_icon="📊")

# --- 2. DATA KONTAK (CONTACT PERSON) ---
# Format: ("Nama / Jabatan", "Nomor Telepon")
DATA_CONTACT = {
    "NKL": [
        ("Putu IC", "087850110155"),
        ("Priyadi IC", "087761390987")
    ],
    "REPORTING": [
        ("Muklis IC", "081934327289"),
        ("Proforma - Ari IC", "081353446516"),
        ("NRB - Yani IC", "087760346299"),
        ("BPB/TAT - Tulasi IC", "081805347302")
    ],
    "RUSAK": [
        ("Putu IC", "087850110155"),
        ("Dwi IC", "083114444424"),
        ("Grand IC", "087725860048")
    ]
}

# --- 3. DATA AKUN & FOLDER ---
AKUN_DIVISI = {
    "REPORTING": {
        "username": "admin_rep",
        "password": "rep123",
        "folder_name": "Reporting",
        "nama_lengkap": "Divisi Reporting / Intransit"
    },
    "NKL": {
        "username": "admin_nkl",
        "password": "nkl123",
        "folder_name": "NKL",
        "nama_lengkap": "Divisi NKL"
    },
    "RUSAK": {
        "username": "admin_rusak",
        "password": "rusak123",
        "folder_name": "BarangRusak",
        "nama_lengkap": "Divisi Barang Rusak"
    }
}

# --- 4. KONEKSI CLOUDINARY ---
def init_cloudinary():
    if "cloudinary" not in st.secrets:
        st.error("⚠️ Kunci Cloudinary belum dipasang!")
        st.stop()
        
    cloudinary.config( 
        cloud_name = st.secrets["cloudinary"]["cloud_name"], 
        api_key = st.secrets["cloudinary"]["api_key"], 
        api_secret = st.secrets["cloudinary"]["api_secret"],
        secure = True
    )

def upload_file(file_upload, nama_folder):
    public_id_path = f"{nama_folder}/{file_upload.name}"
    res = cloudinary.uploader.upload(
        file_upload, 
        resource_type = "raw", 
        public_id = public_id_path,
        overwrite = True
    )
    return res

def hapus_file(public_id):
    try:
        cloudinary.api.delete_resources([public_id], resource_type="raw", type="upload")
        return True
    except Exception as e:
        st.error(f"Gagal menghapus: {e}")
        return False

# --- 5. FUNGSI MAGIC COLOR ---
def get_excel_styles(file_content, sheet_name, header_row, df_original):
    try:
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb[sheet_name]
        
        # Buat DataFrame style kosong dengan struktur sama persis dg data
        styles = pd.DataFrame("", index=df_original.index, columns=df_original.columns)
        
        start_row_excel = header_row + 1 
        
        for r_idx in range(len(df_original)):
            current_excel_row = start_row_excel + r_idx
            row_cells = ws[current_excel_row]
            
            for c_idx in range(len(df_original.columns)):
                if c_idx < len(row_cells):
                    cell = row_cells[c_idx]
                    fill = cell.fill
                    
                    if fill and fill.patternType == 'solid':
                        fg = fill.start_color
                        if fg.type == 'rgb':
                            hex_code = fg.rgb
                            if len(str(hex_code)) > 6:
                                hex_code = str(hex_code)[2:] 
                            color_code = f"#{hex_code}"
                            styles.iat[r_idx, c_idx] = f'background-color: {color_code}'
        return styles
    except Exception as e:
        return None

# --- 6. FUNGSI TAMPILKAN KONTAK (BARU) ---
def tampilkan_kontak(divisi_key):
    """Menampilkan list kontak dalam Expander"""
    kontak_list = DATA_CONTACT.get(divisi_key, [])
    
    # Expander agar bisa diklik buka/tutup
    with st.expander(f"📞 Contact Person (CP) - Divisi {divisi_key}"):
        st.markdown("Silakan hubungi nama di bawah ini jika ada kendala:")
        
        cols = st.columns(len(kontak_list) if len(kontak_list) <= 4 else 4)
        
        for i, (nama, no_telp) in enumerate(kontak_list):
            # Logika kolom agar rapi (wrapping)
            col_idx = i % 4
            
            # Format WA (Ganti 08... jadi 628...)
            no_wa = "62" + no_telp[1:] if no_telp.startswith("0") else no_telp
            
            with cols[col_idx]:
                st.info(f"**{nama}**\n\n[{no_telp}](https://wa.me/{no_wa})")

# --- 7. TAMPILAN PER DIVISI ---
def tampilkan_tab_divisi(nama_divisi, folder_target, semua_files, kode_kontak):
    # 1. Tampilkan Kontak Dulu
    tampilkan_kontak(kode_kontak)
    
    st.divider()
    st.subheader(f"📂 Data: {nama_divisi}")
    
    prefix_folder = folder_target + "/"
    files_divisi = [f for f in semua_files if f['public_id'].startswith(prefix_folder) and f['public_id'].endswith('.xlsx')]
    
    if not files_divisi:
        st.info(f"Folder '{folder_target}' kosong.")
    else:
        dict_files = {}
        for f in files_divisi:
            nama_bersih = f['public_id'].replace(prefix_folder, "")
            dict_files[nama_bersih] = f['secure_url']
        
        pilihan_file = st.selectbox(f"1. Pilih File Excel:", list(dict_files.keys()), key=f"sel_file_{folder_target}")
        
        if pilihan_file:
            url_file = dict_files[pilihan_file]
            
            try:
                response = requests.get(url_file)
                file_content = io.BytesIO(response.content)
                
                xls = pd.ExcelFile(file_content)
                daftar_sheet = xls.sheet_names
                
                st.write("---")
                col1, col2, col3 = st.columns([2, 1, 2])
                
                with col1:
                    sheet_terpilih = st.selectbox("2. Pilih Sheet:", daftar_sheet, key=f"sheet_{folder_target}")
                
                with col2:
                    header_row = st.number_input("3. Header Baris ke-", min_value=1, value=1, key=f"head_{folder_target}")
                
                with col3:
                    cari = st.text_input("🔍 Cari Data:", key=f"search_{folder_target}")

                df = pd.read_excel(file_content, sheet_name=sheet_terpilih, header=header_row - 1)
                
                if cari:
                    mask = df.astype(str).apply(lambda x: x.str.contains(cari, case=False, na=False)).any(axis=1)
                    df_tampil = df[mask]
                else:
                    df_tampil = df

                tampilkan_polos = True
                if len(df_tampil) < 3000:
                    with st.spinner("Mencocokkan warna sel..."):
                        style_matrix = get_excel_styles(file_content, sheet_terpilih, header_row, df)
                        if style_matrix is not None:
                            try:
                                style_tampil = style_matrix.loc[df_tampil.index]
                                st.success(f"Menampilkan Sheet: **{sheet_terpilih}** (Warna Aktif 🎨)")
                                st.dataframe(
                                    df_tampil.style.apply(lambda x: style_tampil, axis=None), 
                                    use_container_width=True, height=600
                                )
                                tampilkan_polos = False
                            except:
                                pass
                
                if tampilkan_polos:
                    st.success(f"Menampilkan Sheet: **{sheet_terpilih}**")
                    st.dataframe(df_tampil, use_container_width=True, height=600)
                
                st.caption(f"Total: {len(df_tampil)} Baris")
                
            except Exception as e:
                st.error("Gagal membaca file.")
                st.error(f"Error: {e}")

# --- 8. PROGRAM UTAMA ---
def main():
    if 'logged_in_divisi' not in st.session_state:
        st.session_state['logged_in_divisi'] = None

    init_cloudinary()

    try:
        raw_files = cloudinary.api.resources(resource_type="raw", type="upload", max_results=500)
        files_list = raw_files.get('resources', [])
    except:
        files_list = []

    # === SIDEBAR ===
    with st.sidebar:
        st.header("🔐 Login Upload")
        
        if st.session_state['logged_in_divisi'] is None:
            target_divisi = st.selectbox("Pilih Divisi:", ["REPORTING", "NKL", "RUSAK"])
            user_input = st.text_input("Username")
            pass_input = st.text_input("Password", type="password")
            
            if st.button("Masuk"):
                akun = AKUN_DIVISI[target_divisi]
                if user_input == akun["username"] and pass_input == akun["password"]:
                    st.session_state['logged_in_divisi'] = target_divisi
                    st.success("Berhasil!")
                    st.rerun()
                else:
                    st.error("Salah password!")
        else:
            divisi_aktif = st.session_state['logged_in_divisi']
            info_akun = AKUN_DIVISI[divisi_aktif]
            folder_aktif = info_akun['folder_name']
            
            st.success(f"Login: {info_akun['nama_lengkap']}")
            
            st.subheader("📤 Upload File")
            uploaded_file = st.file_uploader("Pilih Excel", type=['xlsx'])
            if uploaded_file is not None:
                if st.button(f"Upload ke {divisi_aktif}"):
                    with st.spinner("Mengirim..."):
                        try:
                            upload_file(uploaded_file, folder_aktif)
                            st.success("Sukses upload!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
            
            st.markdown("---")
            st.subheader("🗑️ Hapus File")
            prefix_folder = folder_aktif + "/"
            file_milik_divisi = [f for f in files_list if f['public_id'].startswith(prefix_folder)]
            
            if not file_milik_divisi:
                st.caption("Tidak ada file.")
            else:
                opsi_hapus = {f['public_id'].replace(prefix_folder, ""): f['public_id'] for f in file_milik_divisi}
                file_to_delete = st.selectbox("Pilih file:", list(opsi_hapus.keys()))
                if st.button(f"❌ Hapus {file_to_delete}"):
                     with st.spinner("Menghapus..."):
                         if hapus_file(opsi_hapus[file_to_delete]):
                             st.success("Terhapus.")
                             st.rerun()
            
            st.markdown("---")
            if st.button("🚪 Log Out"):
                st.session_state['logged_in_divisi'] = None
                st.rerun()

    # === HALAMAN UTAMA ===
    st.title("📊 Monitoring IC Bali")
    
    tab_rep, tab_nkl, tab_rusak = st.tabs(["🚛 Reporting", "📉 NKL", "⚠️ Barang Rusak"])

    with tab_rep:
        # Panggil fungsi tab dengan tambahan kode kontak "REPORTING"
        tampilkan_tab_divisi("Reporting", AKUN_DIVISI["REPORTING"]["folder_name"], files_list, "REPORTING")

    with tab_nkl:
        # Panggil fungsi tab dengan tambahan kode kontak "NKL"
        tampilkan_tab_divisi("NKL", AKUN_DIVISI["NKL"]["folder_name"], files_list, "NKL")

    with tab_rusak:
        # Panggil fungsi tab dengan tambahan kode kontak "RUSAK"
        tampilkan_tab_divisi("Barang Rusak", AKUN_DIVISI["RUSAK"]["folder_name"], files_list, "RUSAK")

if __name__ == "__main__":
    main()